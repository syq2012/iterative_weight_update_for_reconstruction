
import matplotlib.colors as mcolors
colors = list(mcolors.TABLEAU_COLORS)


# import functions

import scipy.io as sio

import numpy as np
import scipy

import torch.nn as nn
import torch.utils.data as torchdata
import torch

import importlib
import matplotlib.pyplot as plt
from mpl_toolkits import mplot3d

from sklearn.decomposition import PCA
from sklearn.decomposition import KernelPCA

from numpy import random
from copy import deepcopy
from openpyxl import load_workbook
from random import sample
from sklearn.feature_selection import mutual_info_regression
from sklearn.linear_model import LinearRegression
# own codes 
from RNAupdate import dataset_gen
from RNAupdate import sudo_algo2
from RNAupdate import encoder
from RNAupdate import helper
from RNAupdate import novo_weight




# set figure font sizse 
SMALL_SIZE = 10
plt.rc('font', size=SMALL_SIZE)          # controls default text sizes
plt.rc('axes', titlesize=SMALL_SIZE)     # fontsize of the axes title
plt.rc('axes', labelsize=SMALL_SIZE)    # fontsize of the x and y labels
plt.rc('xtick', labelsize=SMALL_SIZE)    # fontsize of the tick labels
plt.rc('ytick', labelsize=SMALL_SIZE)    # fontsize of the tick labels
plt.rc('legend', fontsize=SMALL_SIZE)    # legend fontsize
# plt.rc('figure', titlesize=SMALL_SIZE)

# some helper function

def reload():
    importlib.reload(encoder)
    importlib.reload(sudo_algo2)
    importlib.reload(dataset_gen)
    importlib.reload(helper)


# reconstruct true average recoverd in CZViz
# define functions for comparing our result to others

def transform_data(x):
    return np.log2(x+10**-4)-np.log2(11*10**-5)

def invert_transform(y):
    return 2**(y+np.log2(11*10**-5))-10**-4 

path = './Circadian-zonation/'
dic_itz = {}
dic_itz_raw = {}
dic_struc = {'rep1': ['00A','06A','12A','18A'], 'rep2': ['00B','06B','12B','18B'], 'rep3': ['00C',None,'12C',None]}
for key, val in dic_struc.items():
    for x in val:
        if x is not None:
#             colab_path = '/content/drive/MyDrive/Colab Notebooks/Circadian-zonation/'
            load_path = path + 'Datasets/Profiles/ZT'+x+'.mat'
            mat = scipy.io.loadmat(load_path)
        for name, data, SD in zip(mat['all_genes'], mat['MeanGeneExp'], mat['SE']):
            if name[0][0] not in dic_itz_raw:
                dic_itz_raw[name[0][0]] = {'rep1' : np.array([]), 'rep1_std' :np.array([]), 'rep2' : np.array([]), 'rep2_std' : np.array([]), 'rep3' : np.array([]), 'rep3_std' :  np.array([])}
                dic_itz[name[0][0]] = {'rep1' : np.array([]), 'rep1_std' :np.array([]), 'rep2' : np.array([]), 'rep2_std' : np.array([]), 'rep3' : np.array([]), 'rep3_std' :  np.array([])}
            if x is None:
                data = [np.nan]*8
                SD = [np.nan]*8
            if len(dic_itz_raw[name[0][0]][key])>0:
                dic_itz_raw[name[0][0]][key] = np.vstack( (dic_itz_raw[name[0][0]][key],np.array(data) ))
                dic_itz_raw[name[0][0]][key+'_std']= np.vstack((dic_itz_raw[name[0][0]][key+'_std'],np.array(SD)))
                dic_itz[name[0][0]][key]= np.vstack( (dic_itz[name[0][0]][key],transform_data(np.array(data))))
                dic_itz[name[0][0]][key+'_std']= np.vstack( (dic_itz[name[0][0]][key+'_std'],transform_data(np.array(SD))))
            else:
                dic_itz_raw[name[0][0]][key] = np.array(data) 
                dic_itz_raw[name[0][0]][key+'_std']= np.array(SD)
                dic_itz[name[0][0]][key]= transform_data(np.array(data))
                dic_itz[name[0][0]][key+'_std']= transform_data(np.array(SD))  
        
#take transpose everywhere
for key in dic_itz:
    for key2 in ['rep1' , 'rep1_std', 'rep2', 'rep2_std', 'rep3', 'rep3_std']:
        dic_itz[key][key2] = dic_itz[key][key2].T
        dic_itz_raw[key][key2] = dic_itz_raw[key][key2].T
        

l_names = list(dic_itz.keys())
#compute list of variance per time condition and per zone condition and then average
l_var = np.array([ np.mean(np.nanvar([dic_itz[gene_name]['rep1'], dic_itz[gene_name]['rep2'],dic_itz[gene_name]['rep3']], axis = 0))/np.nanvar(np.vstack((dic_itz[gene_name]['rep1'], dic_itz[gene_name]['rep2'],dic_itz[gene_name]['rep3']))) for gene_name in l_names])
l_var = np.array([x if not np.isnan(x) else 10**-10 for x in l_var ])
l_exp_log = [invert_transform(np.nanmax(np.vstack((dic_itz[gene_name]['rep1'], dic_itz[gene_name]['rep2'],dic_itz[gene_name]['rep3']))))  for gene_name in l_names]
l_exp = [np.nanmax(np.vstack((dic_itz[gene_name]['rep1'], dic_itz[gene_name]['rep2'],dic_itz[gene_name]['rep3']))) for gene_name in l_names]

def plot_temp(dic, name_gene, reconst):
    f, (ax1) = plt.subplots(1, 1)
    mean = np.zeros(4)
    res = 0
    order = [2, 0, 1, 3]
    for x in range(8):
        
        avg = np.nanmean((dic[name_gene]['rep1'][x,:], dic[name_gene]['rep2'][x,:],dic[name_gene]['rep3'][x,:]), axis = 0)
#         print(len(avg))
#         avg = helper.shift(avg, i)
        std = np.nanstd((dic[name_gene]['rep1'][x,:], dic[name_gene]['rep2'][x,:],dic[name_gene]['rep3'][x,:]), axis = 0)
        ax1.plot(np.linspace(0,18,4, endpoint = True), avg[order], label = ''+str(int((x+1))), lw = 2)
        ax1.fill_between(np.linspace(0,18,4, endpoint = True), (avg+std)[order], (avg-std)[order], alpha=0.25)
        mean += avg
    if reconst != None:
        ax1.plot(np.linspace(0,18,4, endpoint = True), reconst, label = 'reconst', color = 'k', lw = 2)
    ax1.plot(np.linspace(0,18,4, endpoint = True), (mean * 1/8)[order], label = 'mean', color = 'b', lw = 2)
    box = ax1.get_position()
    ax1.set_xticks(np.linspace(0,18,4, endpoint = True), np.array(order) * 6)
#     ax1.set_position([box.x0, box.y0, box.width * 0.8, box.height])
    # Put a legend to the right of the current axis
    ax1.legend(loc='center left', bbox_to_anchor=(1, 0.5))
    #ax2.set_xlim([-0.5,18.5])
    ax1.set_xlim([-0.5,24])
    #ax2.set_xticks([0,6,12,18])
    ax1.set_xticks([0,6,12,18,24])
    ax1.set_xlabel('time')
    ax1.set_ylabel('Expression (fraction of total UMI)')
    plt.show() 
    return res


dic_itz_clean = {}
for name in all_gene:
    if 'mup' not in name and 'pisd' not in name:
        dic_itz_clean[name] = dic_itz[name]
l_names = list(dic_itz_clean.keys())


# import seaborn as sn
def plot_cir(dic, name_gene, reconst, include_plot):
#     color1 = sn.color_palette("husl", 4) 
#     color2 = sn.color_palette("GnBu_d",8)
#     color3 = sn.color_palette("husl", 24)
#     color4 = sn.color_palette("dark", 1)
    if include_plot:
        f, (ax1) = plt.subplots(1, 1)
    mean = np.zeros(8)
    for t in range(4):
        avg = np.nanmean((dic[name_gene]['rep1'][:,t],dic[name_gene]['rep2'][:,t],dic[name_gene]['rep3'][:,t]), axis = 0)
        std = np.nanstd((dic[name_gene]['rep1'][:,t],dic[name_gene]['rep2'][:,t],dic[name_gene]['rep3'][:,t]), axis = 0)
        if include_plot:
            ax1.plot(np.linspace(1,8,8, endpoint = True), avg, label = 'ZT'+str(int(t*6)), lw = 2)
            ax1.fill_between(np.linspace(1,8,8, endpoint = True), avg+std, avg-std, alpha=0.25)
        mean += avg
    mean = mean * 1/4
    if include_plot:
        ax1.plot(np.linspace(1,8,8, endpoint = True), reconst, label = 'reconst', color = 'k', lw = 2)
        ax1.plot(np.linspace(1,8,8, endpoint = True), mean, label = 'mean', color = 'b', lw = 2)
    
    res = np.corrcoef(reconst,mean)[0][1]
#     print(res)
    if include_plot:
        box = ax1.get_position()
        ax1.set_position([box.x0, box.y0, box.width * 0.8, box.height])
        # Put a legend to the right of the current axis
        ax1.legend(loc='center left', bbox_to_anchor=(1, 0.5))
        ax1.set_xlim([0.25,8.25])
        ax1.set_xticks([1,2,3,4,5,6,7,8])
        ax1.set_xlabel("Layer", fontsize=15)
        ax1.set_ylabel("Expression (fraction of total UMI)", fontsize=15)


        plt.show() 
    return res, mean

# plot mean gene expression for each time
def plot_cir_t(dic, name_gene, reconst, include_plot, title):
    if include_plot:
        f, (ax1) = plt.subplots(1, 1)
    res = []
    for t in range(4):
        
        avg = np.nanmean((dic[name_gene]['rep1'][:,t],dic[name_gene]['rep2'][:,t],dic[name_gene]['rep3'][:,t]), axis = 0)
        std = np.nanstd((dic[name_gene]['rep1'][:,t],dic[name_gene]['rep2'][:,t],dic[name_gene]['rep3'][:,t]), axis = 0)
        if include_plot:
            color = next(ax1._get_lines.prop_cycler)['color']
            ax1.plot(np.linspace(1,8,8, endpoint = True), avg, label = 'ZT'+str(int(t*6)), lw = 2, color = color)
            ax1.fill_between(np.linspace(1,8,8, endpoint = True), avg+std, avg-std, alpha=0.25, color = color)
            ax1.plot(np.linspace(1,8,8, endpoint = True), reconst[t], '--',label = 'reconst' + 'ZT'+str(int(t*6)), lw = 2, color = color)
        res.append(np.corrcoef(reconst[t],avg)[0][1])
        print(res)
    if include_plot:
        box = ax1.get_position()
        ax1.set_position([box.x0, box.y0, box.width * 0.8, box.height])
        # Put a legend to the right of the current axis
        ax1.legend(loc='center left', bbox_to_anchor=(1, 0.5))
        ax1.set_xlim([0.25,8.25])
        ax1.set_xticks([1,2,3,4,5,6,7,8])
        ax1.set_xlabel("Layer", fontsize=15)
        ax1.set_ylabel("Expression (fraction of total UMI)", fontsize=15)

        plt.title(title)
        plt.show() 
    return res

def random_partition(l):
    index = [i for i in range(len(l))] 
    np.random.shuffle(index)
    num_partition = int(len(l)/200)

    part_index = []
    for i in range(num_partition):
        upper = (i + 1) * 200
        if i == num_partition:
            upper = len(index)
        part_index.append(index[i * 200: upper])
    
#     part_index = [index[i * 400: np.min((i + 1) * 400, len(l) - 1)] for i in range(num_partition)]
    return part_index

def get_time_cell(time_ran, n):
    res = np.zeros(n)
    s = 0
    for r in time_ran:
        res[r[0]:r[1]] = 1
        s += r[1] - r[0]
    return res, s   

def get_time_base_data(data_mat, time_ran):
    return np.concatenate([data_mat[:, r[0]:r[1]] for r in time_ran], axis = 1)

# load datas
path = './Circadian-zonation/'
dic_struc = ['00A','06A','12A','18A','00B','06B','12B','18B','00C','12C']

wb = load_workbook(path + '42255_2020_323_MOESM3_ESM.xlsx', read_only=True)
ws = wb['schwartz_weights']
dic_weight = {}
for r in ws:
    cur_k = r[0].value
    if cur_k != 'Category' and cur_k != 'Name':
        cur_val = [float(r[i].value) for i in range(1, len(r))]
    else:
        cur_val = [r[i].value for i in range(1, len(r))]
    dic_weight[cur_k] = cur_val
    
# read original gene labels 
dic_genes = {}
wb = load_workbook(path + '42255_2020_323_MOESM2_ESM.xlsx', read_only=True)
for n in wb.sheetnames[:-1]:
    dic_genes[n] = [r[0].value for r in wb[n]]

cat = dic_weight['Category'][:-1]
z_gene = dic_genes['Z'][1:]
r_gene = dic_genes['R'][1:]
zr_gene = dic_genes['Z+R'][1:]
z_r_gene = dic_genes['ZxR'][1:]
known_gene = dic_genes['Z'][1:] + dic_genes['Z+R'][1:] + dic_genes['ZxR'][1:] + dic_genes['R'][1:]
other_gene = dic_genes['F'][1:]

total_gene = known_gene + other_gene

# store number of each types of genes
l_z = len(z_gene)
l_zr = len(zr_gene)
l_z_r = len(z_r_gene)
l_r = len(r_gene)
l_k = len(known_gene)
l_o = len(other_gene)



load_path = path + 'Datasets/Profiles/ZT'+'00A'+'.mat'
cur_data = sio.loadmat(load_path)
all_gene = [temp[0].astype(str)[0].astype(str) for temp in cur_data['all_genes']]


unlable_gene = [g for g in all_gene if not g in total_gene and g != 'pisd']

all_gene = total_gene + unlable_gene
len_k = len(known_gene)
len_z = len(known_gene) - len(r_gene)

l_zonated = ['glul', 'ass1','asl','cyp2f2','cyp1a2','pck1','cyp2e1','cdh1','cyp7a1','acly', 'alb', "oat", 
             "aldob", 'cps1']
central = 'Akr1c6, Alad, Blvrb, C6, Car3, Ccdc107, Cml2, Cyp2c68, Cyp2d9, Cyp3a11, Entpd5, Fmo1, Gsta3, Gstm1, Gstm6, Gstt1, Hpd, Hsd17b10, Inmt, Iqgap2, Mgst1, Nrn1, Pex11a, Pon1, Psmd4, Slc22a1, Tex264'
portal = 'Afm, Aldh1l1, Asl, Ass1, Atp5a1, Atp5g1, C8a, C8b, Ces3b, Cyp2f2, Elovl2, Fads1, Fbp1, Ftcd, Gm2a, Hpx, Hsd17b13, Ifitm3, Igf1, Igfals, Khk, Mug2, Pygl, Sepp1, Serpina1c, Serpina1e, Serpind1, Vtn'
l_central = central.lower().split(', ')
l_portal = portal.lower().split(', ')

path = './Circadian-zonation/'
dic_struc = ['00A','06A','12A','18A','00B','06B','12B','18B','00C','12C']

data_mat_norm_total = helper.load_data_key('mat_norm', dic_struc, path, all_gene)

data_seq_norm_total = helper.load_data_key('seq_data', dic_struc, path,all_gene)

dic_struc = ['00A','06A','12A','18A','00B','06B','12B','18B','00C','12C']
data_mean_total = helper.load_data_key('MeanGeneExp', dic_struc, path,all_gene)


std = [helper.coef_std(data_seq_norm_total[g]) for g in all_gene]

index = np.argsort(np.array(std)[~np.isnan(std)])[-200:]
z_gene_var = [all_gene[i] for i in index if i < len_z]
r_gene_var = [all_gene[i] for i in index if i >= len_z and i < len_k]
other_gene_var = [np.array(all_gene)[~np.isnan(std)][i] for i in index if i > len(known_gene)]


# preprocessing and load data to a dictionary  

norm_to_max = False 

data_0_index = {}
data_total = {}
for g in data_mat_norm_total.keys():
#     print(g)
    temp = data_mat_norm_total[g] + np.array([1e-04] * len(data_mat_norm_total[g]))
    temp = np.log2(np.array(temp))
    temp2 = np.array(helper.shift(temp, g))
    index = [temp2 > 0][0]
    data_0_index[g]= index
    data_total[g] = temp - np.min(temp)

# obtain time label for cells, stored in a dictionary from time to cell index.
dic_struc = {'rep1': ['00A','06A','12A','18A'], 'rep2': ['00B','06B','12B','18B'], 'rep3': ['00C',None,'12C',None]}

dic_pmat = {}
for key, val in dic_struc.items():
    for x in val:
        if x is not None:
#             colab_path = '/content/drive/MyDrive/Colab Notebooks/Circadian-zonation/'
            load_path = path + 'Datasets/Profiles/ZT'+x+'.mat'
            mat = scipy.io.loadmat(load_path)
            pmat = mat['Pmat']
            dic_pmat[x] = pmat
        

time_cell_num = {}
total = 0
for k in dic_pmat:
    num = int(int(k[:2])/6)
    m,n = dic_pmat[k].shape
    if num in time_cell_num:
        time_cell_num[num].append((total, total + m))
    else:
        time_cell_num[num] = [(total, total + m)]
    total += m

# run experiment
cur_list = z_gene_var + r_gene_var + other_gene_var
data_mat = helper.dic_to_array(data_total, [cur_list])
data_dim, cell_dim = data_mat.shape
# 
max_epoch =70
code_dim = 1
    # hidden_layers = [64, 8]
hidden_layers = [64, 8]
step_size = 0.1
num_iter = 20

cur_output_var = sudo_algo2.run_exp(data_mat, max_epoch, code_dim, hidden_layers, step_size, num_iter, 'relu_enc', data_dim, False)


def compute_mean(dic, name_gene):  
    mean = np.zeros(8)
    for t in range(4):
        avg = np.nanmean((dic[name_gene]['rep1'][:,t],dic[name_gene]['rep2'][:,t],dic[name_gene]['rep3'][:,t]), axis = 0)
        std = np.nanstd((dic[name_gene]['rep1'][:,t],dic[name_gene]['rep2'][:,t],dic[name_gene]['rep3'][:,t]), axis = 0)
        mean += avg
    mean = mean * 1/4
    return mean

dic_true_mean = {}
for g in all_gene:
    if 'mup' not in g and 'pisd' not in g:
        dic_true_mean[g] = compute_mean(dic_itz_clean, g)

coef_round = {}
res_plot_dic = {}

flip_list = [4, 7,8,9,10, 13,14,15,16] 

index = [i for i in range(19)]
for i in index:
    print(i)
    code = cur_output_var[-2][i]
    res_plot = {}
    shifted_codes = helper.shift(code, 'none')
    subsets = helper.partial_average_shift_partition(shifted_codes)
    cur_coef = {}
    for g in all_gene:
        if 'mup' not in g and 'pisd' not in g:
            cur = data_mat_norm_total[g]
            temp = helper.partial_average_shift_subset(cur.reshape((-1, 1)), subsets, np.min(cur) - 0.1)

            if i in flip_list:
                temp = np.flip(temp)
            res_plot[g]= temp
    

            cur_coef[g] = np.corrcoef(dic_true_mean[g],transform_data(np.array(temp)))[0][1]

    coef_round[i] = cur_coef
    res_plot_dic[i] = res_plot
    
# plot correlation coef for some rounds
index = [0,1, 3, 5, 7]
# index = [i for i in range(19)]
coef_list2 = {}
for i in index:
    cur_coef = [coef_round[i][g] for g in all_gene if 'mup' not in g and 'pisd' not in g and ~np.isnan(coef_round[i][g])]
    coef_list2[i] = [cur_coef[0:l_z], cur_coef[l_z:l_z + l_zr], cur_coef[l_z + l_zr:l_z + l_zr + l_z_r], cur_coef[l_z + l_zr + l_z_r:l_k], cur_coef[l_k:]]
#     t_datas[n]
counter = 0
for n in index:
    t = counter
    pos = counter - int(len(index)/2)     
    ax.boxplot(coef_list2[n], positions=np.array(range(5))*len(index) +  pos, showfliers = False, patch_artist=True, boxprops=dict(facecolor=colors[t], color=colors[t], alpha=0.6), medianprops=dict(color=colors[t]), capprops=dict(color=colors[t]), whiskerprops=dict(color=colors[t]))
    counter += 1

ticks = ['Z', 'Z+R', 'ZxR', 'R', 'Other']
plt.xticks(range(0, len(ticks) * 5, 5), ticks)
# plt.xlim(-1, len(ticks)*2)

ax.set_ylabel('correlation coefficient')

secx = ax.secondary_xaxis('top')
secx.set_xticks([i - 2 for i in range((len(index)) * 5)])
secx.set_xticklabels(([0, 1, 3, 5, 7]) * 5,  rotation = 45)
secx.set_xlabel('round')

for i in range(3):
    plt.axvspan(-2.5 + i * 10, 2.5 + i * 10, facecolor='0.2', alpha=0.2)

# linear regression
reg_dic_tot = {}
x = np.array([i for i in range(8)]).reshape(-1, 1)
for k in res_plot_dic:
    reg_dic = {}
    res_plot_dic = res_plot_dic[k]
    for n in res_plot_dic:
        res_plot = res_plot_dic[n]
        cur_reg_dic = {}
        for g in all_gene:
            cur_res = []
            cur_lines = res_plot[g]
            for i in range(4):
                temp = LinearRegression().fit(x, transform_data(np.array(cur_lines[i])))
                cur_res.append([temp.coef_[0], temp.score(x, transform_data(np.array(cur_lines[i])))])
            cur_reg_dic[g] = cur_res
        reg_dic[n] = cur_reg_dic
    reg_dic_tot[k] = reg_dic

slope_tot = {}
score_tot = {}
for k in reg_dic_tot:
    reg_dic = reg_dic_tot[k]
    slope = {}
    score = {}
    for i in reg_dic:
        cur_reg_dic = reg_dic[i]
        cur_d = np.array([np.mean(cur_reg_dic[g], axis = 0) for g in all_gene])
        slope[i] = cur_d[:, 0]
        score[i] = cur_d[:, 1]
    slope_tot[k] = slope
    score_tot[k] = score

# save output
slope = {}
for k in slope_tot:
    slope[k] = [slope_tot[k][c].tolist() for c in slope_tot[k]]
with open('./output/score.txt', 'w') as convert_file:
     convert_file.write(json.dumps(score))

score = {}
for k in score_tot:
    score[k] = [score_tot[k][c].tolist() for c in score_tot[k]]
with open('./output/slope.txt', 'w') as convert_file:
     convert_file.write(json.dumps(slope))


# plot the output
labels = ['Z', 'Z + R', 'Z x R', 'other']
part = [0, 1000, l_z + l_zr, l_z + l_zr + l_z_r, len(all_gene)]
for k in slope_5:
    slope = slope_tot[k]
#     print(slope.keys())
    sign = np.sign(slope[4])
    for i in slope:
        slope[i] = slope[i] * sign
def reg_filter(score, slope, sc):
    return (score >= sc) * (slope >= 0.01)
index = [0, 1, 5, 10, 12]
sc = 0.015
for t in range(4):
    res = np.array([[np.sum(reg_filter(slope_5[k][i], score_5[k][i], sc)[part[t]:part[t + 1]])/(np.sum((score_5[k][i] >= sc)[part[t]:part[t + 1]])) for i in index] for k in slope_5])
    plt.errorbar([i for i in range(5)], np.mean(res, axis = 0), np.std(res,axis = 0), capsize=5, label = labels[t])

plt.legend()
ticks = [0, 1, 5, 10, 12]
plt.xticks(range(len(ticks)), ticks)
plt.xlabel('num round')
plt.ylim(0, 1)
plt.ylabel('% Z gene')


# plot some genes 
pl_genes = ['glul','ass1', 'dbp', 'pck1', 'elovl3', 'cyp7a1', 'arntl']
count = 0
index = [7]
for i in index:
    print(i)
    code = cur_output_var2[-2][i]
    res_plot = {}
    shifted_codes = helper.shift(code, 'none')
    subsets = helper.partial_average_shift_partition(shifted_codes)
    for t in range(4):
        code_t = get_time_base_data(np.array(shifted_codes).T, time_cell_num[t])
        for g in pl_genes:
            cur = get_time_base_data(data_mat_norm_total[g].reshape((1, -1)), time_cell_num[t])[0]
            temp = helper.partial_average_shift(cur, code_t[0], np.min(cur) - 0.1)
            if i in [7]:
                temp = np.flip(temp)

            if g in res_plot:
                res_plot[g].append(temp)
            else:
                res_plot[g] = [temp]
    for g in pl_genes:
        res = plot_cir_t(dic_itz_clean, g, transform_data(np.array(res_plot[g])), True, g)
        
# novosparc
var_gene_dic500 = read_dic('hvg_dic500.txt')
var_genes = var_gene_dic500['Z'] + var_gene_dic500['R'] + var_gene_dic500['Other']
l_z500 = len(var_gene_dic500['Z'])

# read umi data

mypath = './Code-to-upload/data/liver_umi/'
onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]

ads = {}
for f in onlyfiles:
    cur_a = ad.read_text(f[:-3])
    cur_a.obs_names_make_unique()
    i = f.index('ZT')
    name = f[i+2:i + 5]
    ads[name] = cur_a


# get full data matrix as annadata and normalize
file_l = dic_struc
data_umi = ad.concat([ads[f] for f in file_l], axis = 1)


def find_coef(tissue_dic, gene_list, data_dic, dim, flip_list):
    res_coef_list = []
    for t in tissue_dic:
        cell_map = tissue_dic[t].gw * dim
        cur_coef = []
        #     t = 0
        for g in gene_list:
            if 'mup' not in g:
                reconst = data_dic[g][subset_cell].dot(cell_map)
                if t in flip_list:
                    temp = transform_data(reconst)
                else:
                    temp = np.flip(transform_data(reconst))
                temp2 = [np.mean(temp[int(i * (dim/8)): int((i + 1)* (dim/8))]) for i in range(8)]

                coef, mean = plot_cir(dic_itz_clean, g, temp2, False)
            cur_coef.append(coef)
        res_coef_list.append(cur_coef)
    return res_coef_list

def subsample_cell(num_cell, num_sample):
    return sample(range(num_cell), num_sample)



subset_cell = subsample_cell(data_umi.n_vars, 8000)

# train on all hvg
filtered_data = data_umi[var_genes, subset_cell].copy().T

sc.pp.normalize_per_cell(filtered_data)
sc.pp.log1p(filtered_data)
print(filtered_data.X.shape)

l_v = len(var_genes)
init_weight = np.array([1/l_v] * l_v)
dim = 8
n_s = 5
n_t = 4
eps = 5e-3
num_itr = 10
lamb = 0.5
update_tiss = novo_weight.multi_weight_novo_param(init_weight, filtered_data, num_itr, dim, n_s, n_t, eps, lamb)

filtered_data = data_umi[var_genes, subset_cell].copy().T
sc.pp.normalize_per_cell(filtered_data)
sc.pp.log1p(filtered_data)
print(filtered_data.X.shape)
# train on zonated hvg
l_v = len(var_genes)
# init_weight = np.array([1/l_v] * l_v)
init_weight = np.array([1/l_z500] * l_z500 + [0] * (l_v - l_z500))
dim = 8
n_s = 5
n_t = 4
eps = 5e-3
num_itr = 1
lamb = 0.5
update_tiss_z = novo_weight.multi_weight_novo_param(init_weight, filtered_data, num_itr, dim, n_s, n_t, eps, lamb)

coef_list = find_coef(update_tiss[0], all_gene, data_mat_norm_total, 8, [])

coef_list_z = find_coef(update_tiss_z[0], all_gene, data_mat_norm_total, 8, [])

# plot the correlation coeffcients

fig, ax = plt.subplots()
index = [0,2, 4, 6, 8]
data = {}
for n in index:
    cur_coef = np.array(coef_list[n])[~np.isnan(coef_list[n])]
    
    data[n] = [cur_coef[0:l_z], cur_coef[l_z:l_z + l_zr], cur_coef[l_z + l_zr:l_z + l_zr + l_z_r], cur_coef[l_z + l_zr + l_z_r:l_k], cur_coef[l_k:]]
#     print(data[n][-1])
for n in [0]:
    cur_coef = np.array(coef_list_z[n])[~np.isnan(coef_list_z[n])]
    data['z'] = [cur_coef[0:l_z], cur_coef[l_z:l_z + l_zr], cur_coef[l_z + l_zr:l_z + l_zr + l_z_r], cur_coef[l_z + l_zr + l_z_r:l_k], cur_coef[l_k:]]
index = index + ['z hvg']
t = 0
for n in data:
    pos = t - int(len(index)/2)
    ax.boxplot(data[n],positions=np.array(range(5))*len(index) +  pos, showfliers = False, patch_artist=True, boxprops=dict(facecolor=colors[t], color=colors[t], alpha=0.6), medianprops=dict(color=colors[t]), capprops=dict(color=colors[t]), whiskerprops=dict(color=colors[t]))
    t += 1

ticks = ['Z', 'Z+R', 'ZxR', 'R', 'Other']
plt.xticks(range(0, len(ticks) * 5, 5), ticks)


secx = ax.secondary_xaxis('top')
secx.set_xticks([i - 3 for i in range(len(index) * 5)])
secx.set_xticklabels((index) * 5,  rotation = 60)
secx.set_xlabel('round')
ax.set_ylabel('correlation coefficient')
# ax.set_xlabel('gene')
for i in range(3):
    plt.axvspan(-3.5 + i * 12, 2.5 + i * 12, facecolor='0.2', alpha=0.2)



